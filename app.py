from stats.html import get_html
from stats.scraper import scraper
import openpyxl
from datetime import datetime
from selenium import webdriver
import time


today = datetime.today().strftime('%Y-%m-%d')

wb = openpyxl.load_workbook('Update.xlsx')
ws = wb['Sheet']

browser = webdriver.Chrome()
browser.get("https://socialblade.com/login")
time.sleep(15)

row = 3
while True:
    try:
        link = ws.cell(row=row, column=1).hyperlink.target
    except:
        break
    username = link.split("/")[-2]
    print(username)
    
    #html = get_html("https://socialblade.com/instagram/user/" + username + "/monthly")
    browser.get("https://socialblade.com/instagram/user/" + username + "/monthly")
    html = browser.page_source
    time.sleep(2)
    
    try:
        a, b, c = scraper(html)
        print(scraper(html))
    except:
        a = "NO DATA"
        b = "NO DATA"
        c = "NO DATA"
        print("NO DATA")
    
    ws.cell(row=row, column=2).value = a
    ws.cell(row=row, column=3).value = b
    ws.cell(row=row, column=4).value = c
    
    row = row + 1
    
row = 3
while True:
    try:
        link = ws.cell(row=row, column=6).hyperlink.target
    except:
        break
    username = link.split("/")[-1]
    print(username)
    
    #html = get_html("https://socialblade.com/twitter/user/" + username + "/monthly")
    browser.get("https://socialblade.com/twitter/user/" + username + "/monthly")
    html = browser.page_source
    time.sleep(2)
    
    try:
        a, b, c = scraper(html)
        print(scraper(html))
    except:
        a = "NO DATA"
        b = "NO DATA"
        c = "NO DATA"
        print("NO DATA")
    
    ws.cell(row=row, column=7).value = a
    ws.cell(row=row, column=8).value = b
    ws.cell(row=row, column=9).value = c
    
    row = row + 1
    
row = 3
while True:
    try:
        link = ws.cell(row=row, column=11).hyperlink.target
    except:
        break
    link = link.split("?")[0]
    username = link.split("@")[-1]
    print(username)
    
    #html = get_html("https://socialblade.com/tiktok/user/" + username + "/monthly")
    browser.get("https://socialblade.com/tiktok/user/" + username + "/monthly")
    html = browser.page_source
    time.sleep(2)
    
    try:
        a, b, c = scraper(html)
        print(scraper(html))
    except:
        a = "NO DATA"
        b = "NO DATA"
        c = "NO DATA"
        print("NO DATA")
    
    ws.cell(row=row, column=12).value = a
    ws.cell(row=row, column=13).value = b
    ws.cell(row=row, column=14).value = c
    
    row = row + 1
    
    browser.quit()
    
    
wb.save(today + ".xlsx")




